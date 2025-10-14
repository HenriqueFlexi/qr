from flask import Flask, request, render_template, jsonify, Response
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
import os
import json
from datetime import datetime

app = Flask(__name__)

EXCEL_FILE = "registros.xlsx"

def load_config():
    with open('config.json', 'r', encoding='utf-8') as f:
        return json.load(f)

def save_config(config):
    with open('config.json', 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

def load_employees():
    with open('employees.json', 'r', encoding='utf-8') as f:
        return json.load(f)

def save_employees(employees):
    with open('employees.json', 'w', encoding='utf-8') as f:
        json.dump(employees, f, ensure_ascii=False, indent=2)

def load_orcamentos():
    if not os.path.exists('orcamentos.json'):
        return []
    with open('orcamentos.json', 'r', encoding='utf-8') as f:
        return json.load(f)

def save_orcamentos(orcamentos):
    with open('orcamentos.json', 'w', encoding='utf-8') as f:
        json.dump(orcamentos, f, ensure_ascii=False, indent=2)

global AREAS, PROJETOS

config = load_config()
AREAS = config['areas']
PROJETOS = config['projetos']

def criar_planilha_se_nao_existir():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"
        headers = ["Data", "ID", "Nome", "Área", "Projeto", "Número Projeto", "Hora Início", "Hora Fim", "Ação"]
        ws.append(headers)
        # Formatar cabeçalhos
        from openpyxl.styles import Font, Alignment
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        # Ajustar larguras das colunas
        column_widths = [10, 5, 30, 20, 15, 20, 12, 12, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

        # Criar aba de gráficos
        ws_chart = wb.create_sheet("Gráficos")
        ws_chart.append(["Área/Projeto", "Horas Trabalhadas", "Horas Orçadas", "Horas Restantes"])
        # Formatar cabeçalhos
        from openpyxl.styles import Font, Alignment
        for col_num, header in enumerate(["Área/Projeto", "Horas Trabalhadas", "Horas Orçadas", "Horas Restantes"], 1):
            cell = ws_chart.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        # Ajustar larguras das colunas
        column_widths = [30, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws_chart.column_dimensions[ws_chart.cell(row=1, column=i).column_letter].width = width

        wb.save(EXCEL_FILE)
    else:
        # Atualizar gráficos sempre que salvar
        atualizar_graficos()

@app.route('/')
def index():
    criar_planilha_se_nao_existir()
    return render_template('index.html')

@app.route('/verificar', methods=['POST'])
def verificar():
    data_req = request.get_json()
    criar_planilha_se_nao_existir()

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Registros"]

    hoje = data_req.get("data")
    idf = data_req.get("id")
    nome = data_req.get("nome")

    aberto = False
    area = ""
    projeto = ""
    numero = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        c_data, c_id, c_nome, c_area, c_proj, c_num, c_inicio, c_fim, c_acao = row
        if (
            c_data == hoje and
            c_id == idf and
            c_nome == nome and
            c_inicio and not c_fim
        ):
            aberto = True
            area = c_area
            projeto = c_proj
            numero = c_num
            break  # Assume only one open at a time

    return jsonify({"aberto": aberto, "area": area, "projeto": projeto, "numeroProjeto": numero})

@app.route('/registrar', methods=['POST'])
def registrar():
    data = request.get_json()
    criar_planilha_se_nao_existir()

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Registros"]

    hoje = data.get("data")
    idf = data.get("id")
    nome = data.get("nome")
    area = data.get("area")
    projeto = data.get("projeto")
    numero = data.get("numeroProjeto")
    hora_atual = datetime.now().strftime("%H:%M")
    acao = ""

    # Busca se já há um registro de entrada sem hora de saída hoje
    linha_encontrada = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        c_data = row[0].value
        c_id = row[1].value
        c_nome = row[2].value
        c_area = row[3].value
        c_proj = row[4].value
        c_num = row[5].value
        c_inicio = row[6].value
        c_fim = row[7].value

        if (
            c_data == hoje and
            c_id == idf and
            c_nome == nome and
            c_area == area and
            c_proj == projeto and
            c_num == numero and
            c_inicio and not c_fim
        ):
            linha_encontrada = row
            break

    if linha_encontrada:
        # Registrar como saída
        linha_encontrada[7].value = hora_atual
        linha_encontrada[8].value = "saída"
        acao = "saída"
    else:
        # Novo registro de entrada
        nova_linha = [
            hoje, idf, nome, area, projeto, numero, hora_atual, "", "entrada"
        ]
        ws.append(nova_linha)
        acao = "entrada"

    try:
        wb.save(EXCEL_FILE)
        atualizar_graficos()
        return jsonify({"status": "ok", "acao": acao})
    except PermissionError:
        return jsonify({"status": "error", "message": "Erro ao salvar no Excel. Verifique se o arquivo está aberto."})

@app.route('/static/config.js')
def config_js():
    import json
    areas_json = json.dumps(AREAS)
    projetos_json = json.dumps(PROJETOS)
    js_code = f"""
    var AREAS = {areas_json};
    var PROJETOS = {projetos_json};
    """
    return Response(js_code, mimetype='application/javascript')

@app.route('/api/employees')
def api_employees():
    employees = load_employees()
    return jsonify(employees)

@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    # Simple hardcoded credentials for demo
    if username == 'admin' and password == 'admin':
        return jsonify({"status": "ok"})
    return jsonify({"status": "error", "message": "Credenciais inválidas"})

@app.route('/admin')
def admin():
    return render_template('admin.html')

@app.route('/admin/add_area', methods=['POST'])
def add_area():
    data = request.get_json()
    area = data.get('area')
    if area and area not in AREAS:
        config['areas'].append(area)
        save_config(config)
        AREAS[:] = config['areas']
        return jsonify({"status": "ok"})
    return jsonify({"status": "error", "message": "Área já existe ou inválida"})

@app.route('/admin/delete_area', methods=['POST'])
def delete_area():
    data = request.get_json()
    area = data.get('area')
    if area in AREAS:
        config['areas'].remove(area)
        save_config(config)
        AREAS[:] = config['areas']
        return jsonify({"status": "ok"})
    return jsonify({"status": "error", "message": "Área não encontrada"})

@app.route('/admin/add_projeto', methods=['POST'])
def add_projeto():
    data = request.get_json()
    projeto = data.get('projeto')
    if projeto and projeto not in PROJETOS:
        config['projetos'].append(projeto)
        save_config(config)
        PROJETOS[:] = config['projetos']
        return jsonify({"status": "ok"})
    return jsonify({"status": "error", "message": "Projeto já existe ou inválido"})

@app.route('/admin/delete_projeto', methods=['POST'])
def delete_projeto():
    data = request.get_json()
    projeto = data.get('projeto')
    if projeto in PROJETOS:
        config['projetos'].remove(projeto)
        save_config(config)
        PROJETOS[:] = config['projetos']
        return jsonify({"status": "ok"})
    return jsonify({"status": "error", "message": "Projeto não encontrado"})

@app.route('/admin/add_employee', methods=['POST'])
def add_employee():
    data = request.get_json()
    id = data.get('id')
    nome = data.get('nome')
    employees = load_employees()
    if id and nome and not any(e['id'] == id for e in employees):
        employees.append({"id": id, "nome": nome})
        save_employees(employees)
        return jsonify({"status": "ok"})
    return jsonify({"status": "error", "message": "Funcionário já existe ou dados inválidos"})

@app.route('/admin/add_orcamento', methods=['POST'])
def add_orcamento():
    data = request.get_json()
    area = data.get('area')
    projeto = data.get('projeto')
    numeroProjeto = data.get('numeroProjeto')
    horasOrcadas = data.get('horasOrcadas')
    orcamentos = load_orcamentos()
    # Check if already exists
    for o in orcamentos:
        if o['area'] == area and o['projeto'] == projeto and o['numeroProjeto'] == numeroProjeto:
            o['horasOrcadas'] = horasOrcadas
            save_orcamentos(orcamentos)
            atualizar_graficos()
            return jsonify({"status": "ok"})
    orcamentos.append({"area": area, "projeto": projeto, "numeroProjeto": numeroProjeto, "horasOrcadas": horasOrcadas})
    save_orcamentos(orcamentos)
    atualizar_graficos()
    return jsonify({"status": "ok"})

@app.route('/api/orcamentos')
def api_orcamentos():
    orcamentos = load_orcamentos()
    return jsonify(orcamentos)

@app.route('/qrcodes')
def qrcodes():
    return render_template('qrcodes.html')

def atualizar_graficos():
    wb = load_workbook(EXCEL_FILE)
    if "Gráficos" not in wb.sheetnames:
        ws_chart = wb.create_sheet("Gráficos")
        ws_chart.append(["Área/Projeto", "Horas Trabalhadas", "Horas Orçadas", "Horas Restantes"])
        # Formatar cabeçalhos
        from openpyxl.styles import Font, Alignment
        for col_num, header in enumerate(["Área/Projeto", "Horas Trabalhadas", "Horas Orçadas", "Horas Restantes"], 1):
            cell = ws_chart.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        # Ajustar larguras das colunas
        column_widths = [30, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws_chart.column_dimensions[ws_chart.cell(row=1, column=i).column_letter].width = width
    else:
        ws_chart = wb["Gráficos"]
        # Clear existing data except header
        for row in ws_chart.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

    orcamentos = load_orcamentos()
    ws_reg = wb["Registros"]

    # Calculate hours worked per area/projeto/numero
    horas_trabalhadas = {}
    for row in ws_reg.iter_rows(min_row=2, values_only=True):
        c_data, c_id, c_nome, c_area, c_proj, c_num, c_inicio, c_fim, c_acao = row
        if c_inicio and c_fim and c_area and c_proj and c_num:
            try:
                inicio = datetime.strptime(c_inicio, "%H:%M")
                fim = datetime.strptime(c_fim, "%H:%M")
                horas = (fim - inicio).total_seconds() / 3600
                key = f"{c_area} - {c_proj} - {c_num}"
                horas_trabalhadas[key] = horas_trabalhadas.get(key, 0) + horas
            except:
                pass

    # Populate chart sheet
    row = 2
    for orc in orcamentos:
        key = f"{orc['area']} - {orc['projeto']} - {orc['numeroProjeto']}"
        trabalhadas = horas_trabalhadas.get(key, 0)
        orcadas = orc['horasOrcadas']
        restantes = max(0, orcadas - trabalhadas)
        ws_chart.cell(row=row, column=1).value = key
        ws_chart.cell(row=row, column=2).value = round(trabalhadas, 2)
        ws_chart.cell(row=row, column=3).value = orcadas
        ws_chart.cell(row=row, column=4).value = round(restantes, 2)
        # Formatar células numéricas
        ws_chart.cell(row=row, column=2).number_format = '0.00'
        ws_chart.cell(row=row, column=3).number_format = '0'
        ws_chart.cell(row=row, column=4).number_format = '0.00'
        row += 1

    # Remove extra rows with None
    max_row = ws_chart.max_row
    while max_row > 1 and all(cell.value is None for cell in ws_chart[max_row]):
        ws_chart.delete_rows(max_row)
        max_row -= 1

    # Create table for filterability
    from openpyxl.worksheet.table import Table, TableStyleInfo
    tab_ref = f"A1:D{row-1}"
    # Remove existing table to avoid name conflicts
    if "DadosGraficos" in ws_chart.tables:
        del ws_chart.tables["DadosGraficos"]
    table = Table(displayName="DadosGraficos", ref=tab_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws_chart.add_table(table)

    # Add or update chart
    # Remove existing charts
    ws_chart._charts = []
    chart = BarChart()
    chart.title = "Horas Trabalhadas vs Orçadas"
    chart.y_axis.title = "Horas"
    chart.x_axis.title = "Área/Projeto"
    data = Reference(ws_chart, min_col=2, min_row=1, max_col=4, max_row=row-1)
    cats = Reference(ws_chart, min_col=1, min_row=2, max_row=row-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    # Note: Chart colors can be customized in Excel after opening the file
    ws_chart.add_chart(chart, "F2")

    wb.save(EXCEL_FILE)

if __name__ == '__main__':
    app.run(debug=True)
